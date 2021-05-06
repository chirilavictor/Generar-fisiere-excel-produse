import openpyxl as xl

wb_inreg = xl.load_workbook('1. Inreg.xlsx')
wb_cana = xl.load_workbook('2. Cani - noi.xlsx')
wb_mousepad = xl.load_workbook('3. Mousepaduri - noi.xlsx')
wb_tcopii = xl.load_workbook('4. Tricouri copii - noi.xlsx')
wb_tbarbati = xl.load_workbook('5. Tricouri barbati - noi.xlsx')
wb_tfemei = xl.load_workbook('6. Tricouri femei - noi.xlsx')
wb_hcopii = xl.load_workbook('7. Hanorace copii - noi.xlsx')
wb_hbarbati = xl.load_workbook('8. Hanorace barbati - noi.xlsx')

sheet_inreg = wb_inreg['Sheet1']

sheet_cana = wb_cana['Sheet1']
sablon_cana = wb_cana['Sablon']

sheet_mousepad = wb_mousepad['Sheet1']
sablon_mousepad = wb_mousepad['Sablon']

sheet_tcopii = wb_tcopii['Sheet1']
sablon_tcopii = wb_tcopii['Sablon']

sheet_tbarbati = wb_tbarbati['Sheet1']
sablon_tbarbati = wb_tbarbati['Sablon']

sheet_tfemei = wb_tfemei['Sheet1']
sablon_tfemei = wb_tfemei['Sablon']

sheet_hcopii = wb_hcopii['Sheet1']
sablon_hcopii = wb_hcopii['Sablon']

sheet_hbarbati = wb_hbarbati['Sheet1']
sablon_hbarbati = wb_hbarbati['Sablon']

sablon_cana_txt = (sablon_cana.cell(2, 1)).value
sablon_mousepad_txt = (sablon_mousepad.cell(2, 1)).value
sablon_tcopii_numefam_txt = (sablon_tcopii.cell(2, 2)).value
sablon_tbarbati_numefam_txt = (sablon_tbarbati.cell(2, 2)).value
sablon_tfemei_numefam_txt = (sablon_tfemei.cell(2, 2)).value
sablon_hcopii_numefam_txt = (sablon_hcopii.cell(2, 2)).value
sablon_hbarbati_numefam_txt = (sablon_hbarbati.cell(2, 2)).value

row_cana = 3
row_mousepad = 3


def nume_produs_simplu(row, sablon_nume, sheet_fisier, row_fisier):
    nume_final = sablon_nume.replace("taguri", (sheet_inreg.cell(row, 1)).value)
    sheet_fisier.cell(row_fisier, 5).value = nume_final


def produs_variabil(row, sablon_numefam, sablon_fisier, sheet_fisier):
    numefam_arr = ((sheet_inreg.cell(row, 1)).value).split(",")
    numefam_final = sablon_numefam.replace("taguri", numefam_arr[0])
    for row_sablon in range(2, sablon_fisier.max_row + 1):
        sablon_numeprodus = (sablon_fisier.cell(row_sablon, 1)).value
        nume_produs_final = sablon_numeprodus.replace("taguri", (sheet_inreg.cell(row, 1)).value)
        #sheet_fisier.cell(sheet_fisier.max_row + 1, 5).value = nume_produs_final
        #sheet_fisier.cell(sheet_fisier.max_row, 10).value = numefam_final
        row_fisier = sheet_fisier.cell(2, 5).value
        sheet_fisier.cell(row_fisier, 5).value = nume_produs_final
        sheet_fisier.cell(row_fisier, 10).value = numefam_final
        row_fisier += 1
        sheet_fisier.cell(2, 5).value = row_fisier

for row in range(2, sheet_inreg.max_row + 1):
    for col in range(2, 9):
        cell_value = (sheet_inreg.cell(row, col)).value
        if cell_value:
            if col == 2:
                # cana_replace = sablon_cana_txt.replace("taguri", (sheet_inreg.cell(row, 1)).value)
                # sheet_cana.cell(row_cana, 5).value = cana_replace
                nume_produs_simplu(row, sablon_cana_txt, sheet_cana, row_cana)
                row_cana += 1
            elif col == 3:
                nume_produs_simplu(row, sablon_mousepad_txt, sheet_mousepad, row_mousepad)
                row_mousepad += 1
            elif col == 4:
                produs_variabil(row, sablon_tcopii_numefam_txt, sablon_tcopii, sheet_tcopii)
                #tcopii_numefam_arr = ((sheet_inreg.cell(row, 1)).value).split(",")
                #tcopii_replace_numefam = sablon_tcopii_numefam_txt.replace("taguri", tcopii_numefam_arr[0])
                #for row_tcopii_sablon in range(2, sablon_tcopii.max_row + 1):
                    #sablon_tcopii_numeprodus_txt = (sablon_tcopii.cell(row_tcopii_sablon, 1)).value
                    #tcopii_replace = sablon_tcopii_numeprodus_txt.replace("taguri", (sheet_inreg.cell(row, 1)).value)
                    #sheet_tcopii.cell(row_tcopii, 5).value = tcopii_replace
                    #sheet_tcopii.cell(row_tcopii, 10).value = tcopii_replace_numefam
                    #row_tcopii += 1
            elif col == 5:
                produs_variabil(row, sablon_tbarbati_numefam_txt, sablon_tbarbati, sheet_tbarbati)
            elif col == 6:
                produs_variabil(row, sablon_tfemei_numefam_txt, sablon_tfemei, sheet_tfemei)
            elif col == 7:
                produs_variabil(row, sablon_hcopii_numefam_txt, sablon_hcopii, sheet_hcopii)
            elif col == 8:
                produs_variabil(row, sablon_hbarbati_numefam_txt, sablon_hbarbati, sheet_hbarbati)


sheet_tcopii.cell(2, 5).value = 3
sheet_tbarbati.cell(2, 5).value = 3
sheet_tfemei.cell(2, 5).value = 3
sheet_hcopii.cell(2, 5).value = 3
sheet_hbarbati.cell(2, 5).value = 3



wb_cana.save('2. Cani - noi.xlsx')
wb_mousepad.save('3. Mousepaduri - noi.xlsx')
wb_tcopii.save('4. Tricouri copii - noi.xlsx')
wb_tbarbati.save('5. Tricouri barbati - noi.xlsx')
wb_tfemei.save('6. Tricouri femei - noi.xlsx')
wb_hcopii.save('7. Hanorace copii - noi.xlsx')
wb_hbarbati.save('8. Hanorace barbati - noi.xlsx')